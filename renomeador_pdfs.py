import os
import re
import shutil
import unicodedata
from pathlib import Path
from datetime import datetime
from tkinter import Tk, Label, END, StringVar, Frame
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
import sys

# --------- PDF deps ----------
try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None
try:
    from PyPDF2 import PdfReader
except Exception:
    PdfReader = None

# --------- Excel deps ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===================== Utilidades =====================

def get_base_dir() -> Path:
    """Retorna a pasta base correta, seja no .py ou no .exe PyInstaller"""
    if getattr(sys, 'frozen', False):  # rodando como .exe
        return Path(sys.executable).parent
    return Path(__file__).parent

def strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def only_digits(s: str) -> str:
    return re.sub(r"\D", "", s)

def read_pdf_text(pdf_path: Path) -> str:
    """Tenta extrair texto do PDF via PyMuPDF; se falhar, tenta PyPDF2."""
    text_parts = []
    if fitz is not None:
        try:
            with fitz.open(pdf_path) as doc:
                for page in doc:
                    text_parts.append(page.get_text("text"))
        except Exception:
            pass
    text = "\n".join(text_parts).strip()
    if not text and PdfReader is not None:
        try:
            reader = PdfReader(str(pdf_path))
            text = "\n".join([p.extract_text() or "" for p in reader.pages])
        except Exception:
            pass
    return text

def validate_cpf(cpf: str) -> bool:
    cpf = only_digits(cpf)
    if len(cpf) != 11 or cpf == cpf[0]*11:
        return False
    soma = sum(int(cpf[i])*(10-i) for i in range(9))
    dv1 = (soma*10) % 11
    if dv1 == 10:
        dv1 = 0
    if dv1 != int(cpf[9]):
        return False
    soma = sum(int(cpf[i])*(11-i) for i in range(10))
    dv2 = (soma*10) % 11
    if dv2 == 10:
        dv2 = 0
    return dv2 == int(cpf[10])


# ===================== Regras de extração =====================

# Aceitar vários tipos de hífen/traço usados em PDFs (U+002D, U+2010..U+2014, U+2212)
_HYPHENS = r"\-\u2010\u2011\u2012\u2013\u2014\u2212"
_SEP = f"[. {_HYPHENS}]"

# CPF com separadores variados (ou ausentes no fallback ANY11)
CPF_REGEX = re.compile(
    rf"\b\d{{3}}(?:{_SEP}?\d{{3}}){{2}}(?:{_SEP}?\d{{2}})\b"
)
CPF_ANY11_REGEX = re.compile(r"\b\d{11}\b")

# Âncora principal do COMPRADOR
ANCHOR_PRINCIPAL = r"DORAVANTE\s+DENOMINADO\(S\)\s+DEVEDOR\(ES\)\s*:"

# âncoras alternativas para fallback
ANCHORS_FALLBACK = [
    r"\bCOMPRADOR(?:ES)?\b",
    r"\bDEVEDOR(?:ES)?\b",
    r"\bPARTE\s+ADQUIRENTE\b",
    r"\bPARTE\s+COMPRADORA\b",
    r"\bADQUIRENTE\b"
]

# Âncoras de término da seção do comprador
END_ANCHORS = [
    r"\bCONSTRUTORA\s+E\s+FIADORA\b",
    r"\bCREDORA\s+FIDUCI[ÁA]RIA\b",
    r"\bVENDEDOR(?:ES)?\b",
    r"\bPARTE\s+VENDEDORA\b",
    r"\bPROMITENTE\s+VENDEDOR(?:ES)?\b",
    r"\bCEDENTE(?:S)?\b",
    r"\bTRANSMITENTE(?:S)?\b",
    r"\bCL[AÁ]USULA\b",
    r"\bOBJETO\b",
    r"\bTESTEMUNHA(?:S)?\b"
]

# Rótulo de CPF tolerante: “CPF”, “C P F”, “C.P.F.”
CPF_LABEL_TOLERANT = re.compile(r"\bC\s*\.?\s*P\s*\.?\s*F\b", flags=re.IGNORECASE)

def _first_end_after(start: int, norm_text: str) -> int:
    """Retorna o índice (em norm_text) do primeiro fim de seção após 'start'."""
    end = len(norm_text)
    for pat in END_ANCHORS:
        m = re.search(pat, norm_text[start:], flags=re.IGNORECASE | re.DOTALL)
        if m:
            cand = start + m.start()
            if cand < end:
                end = cand
    return end

def _extract_cpf_by_label_window(section: str, window_ahead: int = 240) -> str | None:
    """Após o rótulo 'CPF', coleta 11 dígitos à frente (ignorando pontuação)."""
    for m in CPF_LABEL_TOLERANT.finditer(section):
        start = m.end()
        window = section[start: start + window_ahead]
        digits = re.findall(r"\d", window)
        if len(digits) >= 11:
            candidate = "".join(digits[:11])
            if validate_cpf(candidate):
                return candidate
    return None

def extract_cpf_first_buyer(text: str) -> tuple[str | None, str]:
    """
    1) Seção ancorada por 'doravante denominado(s) DEVEDOR(ES):'
    2) Busca CPF pós-rótulo; fallback para padrões válidos; último recurso: fullscan.
    """
    if not text or not text.strip():
        return None, "ERRO: PDF sem texto"

    norm = strip_accents(text).upper()

    # 1) tentativa principal
    m_main = re.search(ANCHOR_PRINCIPAL, norm)
    if m_main:
        start_pos = m_main.end()
        end_pos = _first_end_after(start_pos, norm)
        end_pos = max(end_pos, start_pos)
        section = text[start_pos: start_pos + min(25000, end_pos - start_pos or 25000)]

        cpf = _extract_cpf_by_label_window(section)
        if cpf:
            return cpf, "OK"

        m = CPF_REGEX.search(section)
        if m and validate_cpf(only_digits(m.group(0))):
            return only_digits(m.group(0)), "OK"

        m = CPF_ANY11_REGEX.search(section)
        if m and validate_cpf(m.group(0)):
            return m.group(0), "OK"

        return None, "ERRO: CPF não encontrado após âncora principal"

    # 2) fallback: âncoras alternativas
    for rgx in ANCHORS_FALLBACK:
        m_fb = re.search(rgx, norm)
        if not m_fb:
            continue
        start_pos = m_fb.end()
        end_pos = _first_end_after(start_pos, norm)
        end_pos = max(end_pos, start_pos)
        section = text[start_pos: start_pos + min(15000, end_pos - start_pos or 15000)]

        cpf = _extract_cpf_by_label_window(section)
        if cpf:
            return cpf, "⚠️ Âncora principal não encontrada, CPF via fallback"

        m = CPF_REGEX.search(section)
        if m and validate_cpf(only_digits(m.group(0))):
            return only_digits(m.group(0)), "⚠️ Âncora principal não encontrada, CPF via fallback"

        m = CPF_ANY11_REGEX.search(section)
        if m and validate_cpf(m.group(0)):
            return m.group(0), "⚠️ Âncora principal não encontrada, CPF via fallback"

    # 3) fullscan (priorizando ainda o rótulo)
    m = CPF_LABEL_TOLERANT.search(text)
    if m:
        w = text[m.end(): m.end() + 240]
        digits = re.findall(r"\d", w)
        if len(digits) >= 11:
            candidate = "".join(digits[:11])
            if validate_cpf(candidate):
                return candidate, "⚠️ Âncora não encontrada, CPF via fullscan"

    for m in CPF_REGEX.finditer(text):
        cand = only_digits(m.group(0))
        if validate_cpf(cand):
            return cand, "⚠️ Âncora não encontrada, CPF via fullscan"

    for m in CPF_ANY11_REGEX.finditer(text):
        cand = m.group(0)
        if validate_cpf(cand):
            return cand, "⚠️ Âncora não encontrada, CPF via fullscan"

    return None, "ERRO: Nenhum CPF válido encontrado"


# --------- EXTRATOR do número do contrato (ajustado) ---------

# Aceita também Nº / N° / No, com pontuação opcional
_N_LABEL = r"N[\s\.\-º°O]*"

def _take_first_13_digits(s: str) -> str | None:
    """Retorna os primeiros 13 dígitos encontrados na string."""
    ds = re.findall(r"\d", s)
    if len(ds) >= 13:
        return "".join(ds[:13])
    return None

def extract_contract_number(text: str) -> str | None:
    """
    Estratégia:
      A) Procurar linhas que contenham 'CONTRATO'. Tentar na própria linha e na seguinte:
         - Preferir: pegar os PRIMEIROS 13 dígitos após 'Nº/No/N°' (ou após 'CONTRATO' se não houver rótulo).
         - Isso evita capturar a numeração de página ao final da linha.
      B) Fallback curto (OCR 'quebrado'): nas primeiras ~40 linhas, procurar linhas com rótulo de número (Nº/No/N°)
         e extrair os PRIMEIROS 13 dígitos após o rótulo (preferir linhas que também contenham 'CONTR').
    """
    if not text:
        return None

    lines_orig = text.splitlines()
    lines_norm = [strip_accents(l).upper() for l in lines_orig]

    # Linhas com 'CONTRATO'
    candidate_idxs = [i for i, ln in enumerate(lines_norm) if "CONTRATO" in ln]

    # Regex para localizar 'CONTRATO' e depois um rótulo N...
    # (usamos só para achar a região; a extração final é "primeiros 13 dígitos após")
    num_after_contrato_regex = re.compile(
        rf"CONTRATO(?P<after>.*)",  # pega tudo após 'CONTRATO'
        flags=re.IGNORECASE
    )
    nlabel_regex = re.compile(_N_LABEL, flags=re.IGNORECASE)

    def try_window(idx: int) -> str | None:
        # janela: linha do 'CONTRATO' + próxima (para contornar quebras)
        win = lines_orig[idx]
        if idx + 1 < len(lines_orig):
            win += " " + lines_orig[idx + 1]

        m_contrato = num_after_contrato_regex.search(win)
        if not m_contrato:
            return None

        tail = m_contrato.group("after")

        # 1) Se existir rótulo N... dentro da janela, pegar a partir dele
        m_n = nlabel_regex.search(tail)
        if m_n:
            sub = tail[m_n.end():]
            n = _take_first_13_digits(sub)
            if n:
                return n

        # 2) Sem rótulo: pegar a partir do fim de 'CONTRATO'
        n = _take_first_13_digits(tail)
        if n:
            return n

        # 3) Fallback dentro da janela: maior trecho numérico e primeiros 13 dígitos
        candidates = re.findall(rf"[0-9{_HYPHENS}\.\s]{{13,}}", tail)
        if candidates:
            best = max(candidates, key=lambda s: len(re.findall(r"\d", s)))
            n = _take_first_13_digits(best)
            if n:
                return n
        return None

    for idx in candidate_idxs:
        n = try_window(idx)
        if n:
            return n

    # Fallback curto: primeiras ~40 linhas que tenham rótulo N... (preferir com 'CONTR')
    limit = min(40, len(lines_orig))
    for i in range(limit):
        ln = lines_orig[i]
        ln_up = lines_norm[i]
        if nlabel_regex.search(ln_up) and ("CONTR" in ln_up or i < 10):
            # pegar texto após o rótulo e extrair os PRIMEIROS 13 dígitos
            m = nlabel_regex.search(ln)
            sub = ln[m.end():] if m else ln
            n = _take_first_13_digits(sub)
            if n:
                return n

            # fallback: maior bloco numérico na linha e PRIMEIROS 13 dígitos
            candidates = re.findall(rf"[0-9{_HYPHENS}\.\s]{{13,}}", ln)
            if candidates:
                best = max(candidates, key=lambda s: len(re.findall(r"\d", s)))
                n = _take_first_13_digits(best)
                if n:
                    return n

    return None


def extract_nome_until_comma(text: str, anchors_regexes: list[str]) -> str | None:
    raw = text
    up = strip_accents(raw.upper())
    start_pos = None
    for rgx in anchors_regexes:
        m = re.search(rgx, up)
        if m:
            start_pos = m.end()
            break
    if start_pos is None:
        return None
    tail_up = up[start_pos:]
    cut_positions = []
    comma_pos = tail_up.find(',')
    if comma_pos != -1:
        cut_positions.append(comma_pos)
    STOPWORDS_POS_NOME = [
        r"\bCPF\b", r"\bRG\b", r"\bCNH\b", r"\bCTPS\b",
        r"\bFILIA[CÇ][AÃ]O\b", r"\bNASC\w*\b", r"\bNATURAL\b",
        r"\bRESIDENTE\b", r"\bENDERE[CÇ]O\b",
        r"\bCONFORME\b", r"\bREQUERID\w*\b", r"\bREQUERENTE\w*\b",
        r"\bCERTID[ÃA]O\b", r"\bEMITID\w*\b", r"\bEXPEDID\w*\b",
        r"\bPROCESSO\b", r"\bPORTADOR\b", r"\bPORTADORA\b",
        r"\bEM\b", r"\bE\b"
    ]
    for st in STOPWORDS_POS_NOME:
        m = re.search(st, tail_up)
        if m:
            cut_positions.append(m.start())
    if cut_positions:
        cut_at = min([pos for pos in cut_positions if pos >= 0])
        candidate = tail_up[:cut_at]
    else:
        first_line = next((ln.strip() for ln in tail_up.splitlines() if ln.strip()), "")
        candidate = first_line
    candidate = re.sub(r"[^A-Z \n]", " ", candidate)
    candidate = re.sub(r"\s+", " ", candidate).strip()
    return candidate if candidate else None

def extract_oficio_num(text: str) -> str:
    txt = strip_accents(text.upper())
    if re.search(r"\b6[ºO]?\s*OFICIO\b", txt):
        return "6"
    return "5"


# ===================== Casos de renomeação =====================

def rename_contratos(pdf: Path, outdir: Path):
    text = read_pdf_text(pdf)
    if not text:
        return (pdf.name, "", "ERRO: PDF sem texto (digitalizado sem OCR?)")

    cpf, status = extract_cpf_first_buyer(text)
    contrato = extract_contract_number(text)

    if not cpf:
        return (pdf.name, "", status)
    if not contrato:
        return (pdf.name, "", "ERRO: Número de contrato (13 dígitos) não encontrado")

    novo = f"{cpf}_{contrato}.pdf"
    dest = outdir / novo
    i = 1
    while dest.exists():
        dest = outdir / f"{cpf}_{contrato}_{i}.pdf"
        i += 1
    shutil.copy2(pdf, dest)
    return (pdf.name, dest.name, status)


def rename_certidoes_2(pdf: Path, outdir: Path):
    text = read_pdf_text(pdf)
    if not text:
        return (pdf.name, "", "ERRO: PDF sem texto (digitalizado sem OCR?)")
    nome = extract_nome_until_comma(text, [
        r"COM\s+REFERENCIA\s+AO\s+NOME\s+DE"
    ])
    if not nome:
        return (pdf.name, "", "ERRO: Nome não encontrado após 'COM REFERENCIA AO NOME DE' até a vírgula")
    dest = outdir / f"{nome}-2.pdf"
    i = 1
    while dest.exists():
        dest = outdir / f"{nome}-2.{i}.pdf"
        i += 1
    shutil.copy2(pdf, dest)
    return (pdf.name, dest.name, "OK")

def rename_certidoes_5_6(pdf: Path, outdir: Path):
    text = read_pdf_text(pdf)
    if not text:
        return (pdf.name, "", "ERRO: PDF sem texto (digitalizado sem OCR?)")
    nome = extract_nome_until_comma(text, [
        r"NADA\s+CONSTA\s+EM\s+NOME\s+DE",
        r"EM\s+NOME\s+DE",
    ])
    if not nome:
        return (pdf.name, "", "ERRO: Nome não encontrado após 'NADA CONSTA EM NOME DE/EM NOME DE' até a vírgula")
    oficio = extract_oficio_num(text)
    dest = outdir / f"{nome}-{oficio}.pdf"
    i = 1
    while dest.exists():
        dest = outdir / f"{nome}-{oficio}.{i}.pdf"
        i += 1
    shutil.copy2(pdf, dest)
    return (pdf.name, dest.name, "OK")


# ===================== Interface (Tkinter) =====================

PRIMARY = "#007BFF"
PRIMARY_DARK = "#0056b3"
BG = "#F4F8FF"
CARD_BG = "#E8F1FF"
TEXT_DARK = "#1F2937"

APP_VERSION = "0.0.5"

class App:
    def __init__(self, master: Tk):
        self.master = master
        master.title(f"Renomeador de PDFs v{APP_VERSION}")
        master.geometry("900x600")
        master.configure(bg=BG)

        # Título
        Label(master, text=f"📂 Renomeador de PDFs (v{APP_VERSION})",
              bg=BG, fg=PRIMARY, font=("Segoe UI", 20, "bold")).pack(pady=(14, 6))

        Label(master, text="Selecione o tipo de documento",
              bg=BG, fg=TEXT_DARK, font=("Segoe UI", 12)).pack()

        # ===== Cards de modos =====
        self.mode = StringVar(value="contratos")
        self.cards_frame = Frame(master, bg=BG)
        self.cards_frame.pack(pady=12)
        self.cards = {}
        self._create_card("Contratos", "contratos", 0)
        self._create_card("Certidões -2", "certidoes_2", 1)
        self._create_card("Certidões 5/6", "certidoes_5_6", 2)

        # botão executar
        self.run_btn = ttk.Button(master, text="Executar", command=self.run)
        self.run_btn.pack(pady=12)

        Label(master, text="Log:", bg=BG, fg=TEXT_DARK, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=14)

        self.log = ScrolledText(master, height=16, wrap="word", font=("Consolas", 10))
        self.log.pack(fill="both", expand=True, padx=12, pady=(4, 12))
        self.log.configure(state="disabled")

        self.log.tag_config("ok", foreground="green")
        self.log.tag_config("warn", foreground="orange")
        self.log.tag_config("err", foreground="red")
        self.log.tag_config("normal", foreground=TEXT_DARK)

        self.ensure_folders()
        self._apply_card_styles()

    def _create_card(self, title: str, mode_value: str, col: int):
        card = Frame(self.cards_frame, bg=CARD_BG, bd=0, highlightthickness=1, highlightbackground="#D0E2FF")
        lbl = Label(card, text=title, bg=CARD_BG, fg=TEXT_DARK, font=("Segoe UI", 11, "bold"), padx=24, pady=18)
        card.grid(row=0, column=col, padx=10, pady=6, sticky="n")
        lbl.pack()
        card.bind("<Button-1>", lambda e, m=mode_value: self._select_mode(m))
        lbl.bind("<Button-1>", lambda e, m=mode_value: self._select_mode(m))
        self.cards[mode_value] = (card, lbl)
        self.cards_frame.grid_columnconfigure(col, weight=1)

    def _select_mode(self, mode_value: str):
        self.mode.set(mode_value)
        self._apply_card_styles()
        self._log_add(f"ℹ️ Modo selecionado: {mode_value}\n", "normal")

    def _apply_card_styles(self):
        for m, (card, lbl) in self.cards.items():
            if self.mode.get() == m:
                card.configure(bg=PRIMARY)
                lbl.configure(bg=PRIMARY, fg="white")
                card.configure(highlightbackground=PRIMARY_DARK)
            else:
                card.configure(bg=CARD_BG)
                lbl.configure(bg=CARD_BG, fg=TEXT_DARK)
                card.configure(highlightbackground="#D0E2FF")

    def _log_add(self, msg: str, tag: str = "normal"):
        self.log.configure(state="normal")
        self.log.insert(END, msg, tag)
        self.log.see(END)
        self.log.configure(state="disabled")

    def ensure_folders(self):
        base = get_base_dir()
        for p in ["contratos", "certidoes_2", "certidoes_5_6"]:
            (base / "entrada" / p).mkdir(parents=True, exist_ok=True)
            (base / "saida" / p).mkdir(parents=True, exist_ok=True)
        self._log_add("📁 Pastas verificadas/criadas nas pastas 'entrada' e 'saida'.\n", "normal")

    def run(self):
        base = get_base_dir()
        mode = self.mode.get()
        entrada = base / "entrada" / mode
        saida = base / "saida" / mode

        fn = {
            "contratos": rename_contratos,
            "certidoes_2": rename_certidoes_2,
            "certidoes_5_6": rename_certidoes_5_6
        }[mode]

        pdfs = sorted(entrada.glob("*.pdf"))
        if not pdfs:
            self._log_add(f"⚠️ A pasta '{entrada}' está vazia. Nenhum PDF para processar.\n", "warn")
            return

        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        xlsx_path = saida / f"renomeacao_log_{ts}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Log Renomeação"
        headers = ["Arquivo Original", "Nome Novo", "Status"]
        ws.append(headers)

        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DDEBFF")
        header_align = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="D0D7E2")
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        ok = 0
        for pdf in pdfs:
            try:
                orig, novo, status = fn(pdf, saida)
            except Exception as e:
                orig, novo, status = (pdf.name, "", f"ERRO: {e}")
            ws.append([orig, novo, status])
            if status == "OK":
                self._log_add(f"✅ {orig} → {novo}\n", "ok")
                ok += 1
            elif isinstance(status, str) and status.startswith("⚠️"):
                self._log_add(f"{status} | {orig} → {novo}\n", "warn")
            else:
                self._log_add(f"❌ {orig} → {status}\n", "err")

        # ajustar colunas
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                s = str(val) if val is not None else ""
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(80, max_len + 2))

        # bordas & destaque erros
        thin = Side(border_style="thin", color="D0D7E2")
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if col_idx == 3 and isinstance(c.value, str) and c.value.startswith("ERRO"):
                    c.font = Font(color="FF0000")

        wb.save(xlsx_path)
        self._log_add(f"\n📊 Log detalhado salvo em: {xlsx_path}\n", "normal")
        self._log_add(f"\nConcluído. {ok} arquivos renomeados.\n", "ok")


if __name__ == "__main__":
    root = Tk()
    app = App(root)
    root.mainloop()